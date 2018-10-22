from openpyxl import Workbook  # Librería para manipular archivos xlsx
import osarga en memoria
import random
from time import sleep
# PROGRAMA HECHO POR:
# Esteban Emmanuel Brito Borges
# Victor Emmanuel Chan Cocom
# Raúl Armín Novelo Cruz
# David Iván Mendoza Rodríguez
# Fernando Ángel Adrián Herrera Garnica

class OS:
    # Esta clase emula la funcionalidad del sistema operativo
    # al asignar memoria y recursos a las páginas y secciones de cada proceso.
    def __init__(self):
        self._actual_page = 0
        self._s = Workbook()
        self._seg_table = self._s.active
        self._p = Workbook()
        self._pag_table = self._p.active
        self._l = Workbook()
        self._log_memory = self._l.active
        self._ph = Workbook()
        self._phy_memory = self._ph.active

        # Escribir etiquetas de tablas para cada archivo
        # (Segments Table)
        self._seg_table.cell(row=1, column=1, value="Segments A")
        self._seg_table.cell(row=1, column=3, value="Segments B")
        self._seg_table.cell(row=1, column=5, value="Segments C")
        for i in range(3):
            # Se ponen las etiquetas de sección a cada tabla de segmentos
            self._seg_table.cell(row=2, column=1 + 2*i, value="CODE")
            self._seg_table.cell(row=3, column=1 + 2*i, value="VARIABLES")
            self._seg_table.cell(row=4, column=1 + 2*i, value="STACK")
        for i in range(3):
            self._seg_table.cell(row=1, column=2 + 2*i, value="Frame")

        # (Pages table)
        self._pag_table.cell(row=1, column=1, value="Page")
        self._pag_table.cell(row=1, column=2, value="Frame")
        for i in range(18):
            # Se enumeran del 0 al 17 los marcos existentes
            self._pag_table.cell(row=2 + i, column=2, value=i)

        # (Logical Memory)
        self._log_memory.cell(row=1, column=1, value="Frame")
        self._log_memory.cell(row=1, column=2, value="Bit")
        self._log_memory.cell(row=1, column=3, value="State")
        for i in range(18):
            # Se enumeran del 0 al 17 los marcos existentes
            self._log_memory.cell(row=2 + i, column=1, value=i)
            # Se enumeran los bits que señala cada marco
            self._log_memory.cell(row=2 + i, column=2, value=4*i) # (value = 16+4*i con SO emulado)
            # Se inicializa el valor de cada celda (0 es libre, 1 es ocupado)
            self._log_memory.cell(row=2 + i, column=3, value=0)

        # (Physical Memory)
        self._phy_memory.cell(row=1, column=1, value="Bit")
        self._phy_memory.cell(row=1, column=2, value="Content")
        # Para emular SO
        # for i in range(15):
        #   self._phy_memory.cell(row=2 + i, column=2, value=XXXXXX)
        for i in range(18*4):
            # Se enumeran del 0 al 71 (87 con SO emulado) los bits de la memoria física
            self._phy_memory.cell(row=2 + i, column=1, value=i)

    def getPage(self):
        return self._actual_page

    def writeInMemory(self, bits, name, section):
        """Divide en páginas y escribe en memoria el contenido de un proceso"""
        label = "PRO" + name + "-" + section + "-PAGE"  # Algo como PRO1-CODE-PAGE or P3-VARS-PAGE
        page = 0
        while bits != 0:
            if bits >= 4:
                bits_to_write = 4
                bits -= 4
            else:
                bits_to_write = bits
                bits = 0
            # Sobreescribe bits en Memoria Física
            for i in range(bits_to_write):
                self._phy_memory.cell(row=2 + 4*self._actual_page + 4*page + i, column=2, value=label + str(page))
            page += 1

        # Registra la sobreescritura de memoria en las otras tablas
        # (Segments Table)
        if name == "A":
            columna = 2
        elif name == "B":
            columna = 4
        else:  # Es C
            columna = 6

        if section == "CODE":
            fila = 2
        elif section == "VARS":
            fila = 3
        else:  # Es "STACK"
            fila = 4
        self._seg_table.cell(row=fila, column=columna, value=self._actual_page)

        # (Pages Table)
        for i in range(page):
            self._pag_table.cell(row=2 + self._actual_page + i, column=1, value=label + str(i))

        # (Logical Memory)
        for i in range(page):
            self._log_memory.cell(row=2 + self._actual_page + i, column=3, value=1)

        # Actualiza el valor de la página actual a escribir
        self._actual_page += page

    def close(self):
        """Guarda las cambios hechos en Excel en archivos xlsx"""
        self._s.save(os.path.dirname(__file__) + "/segments_table.xlsx")
        self._p.save(os.path.dirname(__file__) + "/pages_table.xlsx")
        self._l.save(os.path.dirname(__file__) + "/logical_memory.xlsx")
        self._ph.save(os.path.dirname(__file__) + "/physical_memory.xlsx")


class Process:
    def __init__(self, letter):
        self._name = letter
        # Cada sección del proceso recibe un número
        # aleatorio de bits que es múltiplo de 4
        self._code_bits = random.randint(1, 2) * 4
        self._variable_bits = random.randint(1, 2) * 4
        self._stack_bits = random.randint(1, 2) * 4

    def printState(self):
        """Imprime el número de bits de cada sección del proceso"""
        print("\nProceso:", self._name)
        print("Seccion | Bits")
        print("Código      ", end="")
        print(self._code_bits)
        print("Variables   ", end="")
        print(self._variable_bits)
        print("Pila        ", end="")
        print(self._stack_bits)

    def getName(self):
        return self._name

    def getCodeBits(self):
        return self._code_bits

    def getVarBits(self):
        return self._variable_bits

    def getStackBits(self):
        return self._stack_bits


if __name__ == "__main__":
    letters = ["A", "B", "C"]

    op_sys = OS()

    # Crear 3 procesos y asignarle una letra como ID
    for i in range(3):
        p = Process(letters[i])
        p.printState()

        # Cada proceso pide al sistema op que cargue en memoria sus secciones
        op_sys.writeInMemory(p.getCodeBits(), p.getName(), "CODE")
        op_sys.writeInMemory(p.getVarBits(), p.getName(), "VARS")
        op_sys.writeInMemory(p.getStackBits(), p.getName(), "STACK")

        sleep(1)
    op_sys.close()
